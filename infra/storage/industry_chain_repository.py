# -*- coding: utf-8 -*-
"""XLSX and signature-cache repository for the AI industry-chain pool."""

from __future__ import annotations

import json
import os
import threading
import uuid
from pathlib import Path

from core.logger import get_logger
from core.observability import emit_structured_log, record_metric
from domains.industry_chain.pool_service import (
    build_ai_industry_chain_context_map,
    build_ai_industry_chain_rows,
    normalize_ai_chain_code,
)

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_PROJECT_PARENT = Path(__file__).resolve().parents[3]

AI_CHAIN_CANDIDATE_PATHS = (
    _PROJECT_PARENT / "产业链投研" / "watchlists" / "AI产业链.xlsx",
    _PROJECT_PARENT / "产业链投研" / "AI产业链.xlsx",
    _PROJECT_ROOT / "data" / "AI产业链.xlsx",
)


def resolve_ai_chain_file() -> Path:
    for candidate in AI_CHAIN_CANDIDATE_PATHS:
        try:
            if candidate.is_file():
                return candidate
        except OSError:
            continue
    return AI_CHAIN_CANDIDATE_PATHS[0]


AI_CHAIN_FILE = resolve_ai_chain_file()
AI_CHAIN_ROWS_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_rows.json"
AI_CHAIN_CODES_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_stock_codes.json"
AI_CHAIN_CONTEXT_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_context_map.json"
_CACHE_LOCK = threading.RLock()
_CACHE_READ_ERRORS = (OSError, TypeError, ValueError, json.JSONDecodeError)
_CACHE_WRITE_ERRORS = (OSError, TypeError, ValueError)
_log = get_logger(__name__)


class IndustryChainRepository:
    """Own all filesystem and workbook I/O for the industry-chain pool."""

    def __init__(
        self,
        *,
        workbook_path: str | Path = AI_CHAIN_FILE,
        rows_cache_path: str | Path = AI_CHAIN_ROWS_CACHE_FILE,
        codes_cache_path: str | Path = AI_CHAIN_CODES_CACHE_FILE,
        context_cache_path: str | Path = AI_CHAIN_CONTEXT_CACHE_FILE,
    ) -> None:
        self.workbook_path = Path(workbook_path)
        self.rows_cache_path = Path(rows_cache_path)
        self.codes_cache_path = Path(codes_cache_path)
        self.context_cache_path = Path(context_cache_path)

    @staticmethod
    def _same_path(left: Path, right: Path) -> bool:
        try:
            return left.resolve() == right.resolve()
        except OSError:
            return left == right

    def _resolve_source_path(self, source_path: str | Path | None = None) -> Path:
        if source_path is not None:
            path = Path(source_path)
            if path.exists():
                return path
            if path.name == "AI产业链.xlsx":
                resolved = resolve_ai_chain_file()
                if resolved.exists():
                    return resolved
            return path
        if self.workbook_path.exists():
            return self.workbook_path
        resolved = resolve_ai_chain_file()
        if resolved.exists():
            return resolved
        return self.workbook_path

    def is_default_source(self, source_path: str | Path) -> bool:
        path = Path(source_path)
        if self._same_path(path, self.workbook_path):
            return True
        for candidate in AI_CHAIN_CANDIDATE_PATHS:
            if self._same_path(path, candidate):
                return True
        return False

    @staticmethod
    def source_signature(source_path: str | Path) -> dict[str, object]:
        path = Path(source_path)
        if not path.exists() and path.name == "AI产业链.xlsx":
            resolved = resolve_ai_chain_file()
            if resolved.exists():
                path = resolved
        stat = path.stat()
        try:
            resolved_path = path.resolve()
        except OSError:
            resolved_path = path
        return {
            "path": str(resolved_path),
            "size": int(stat.st_size),
            "mtime_ns": int(stat.st_mtime_ns),
        }

    @staticmethod
    def source_mtime(source_path: str | Path) -> float:
        path = Path(source_path)
        if not path.exists() and path.name == "AI产业链.xlsx":
            resolved = resolve_ai_chain_file()
            if resolved.exists():
                path = resolved
        try:
            return float(path.stat().st_mtime)
        except (FileNotFoundError, OSError, TypeError, ValueError):
            return 0.0

    def _read_signature_cache(self, cache_path: Path, source_path: Path, payload_key: str):
        signature = self.source_signature(source_path)
        with _CACHE_LOCK:
            try:
                with cache_path.open("r", encoding="utf-8") as handle:
                    payload = json.load(handle)
            except _CACHE_READ_ERRORS as exc:
                _log.warning("[产业链池] 读取签名缓存失败: %s", exc, exc_info=True)
                return None
        if not isinstance(payload, dict) or payload.get("source_signature") != signature:
            return None
        return payload.get(payload_key)

    def _write_signature_cache(self, cache_path: Path, source_path: Path, payload_key: str, value: object) -> None:
        payload = {
            "source_signature": self.source_signature(source_path),
            payload_key: value,
        }
        try:
            with _CACHE_LOCK:
                cache_path.parent.mkdir(parents=True, exist_ok=True)
                temp_path = cache_path.with_name(f"{cache_path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
                try:
                    with temp_path.open("w", encoding="utf-8") as handle:
                        json.dump(payload, handle, ensure_ascii=False)
                        handle.flush()
                        os.fsync(handle.fileno())
                    with temp_path.open("r", encoding="utf-8") as handle:
                        json.load(handle)
                    os.replace(temp_path, cache_path)
                finally:
                    try:
                        temp_path.unlink(missing_ok=True)
                    except OSError:
                        pass
        except _CACHE_WRITE_ERRORS as exc:
            _log.warning("[产业链池] 写入签名缓存失败: %s", exc, exc_info=True)
            emit_structured_log(
                "industry_chain.cache_write.failed",
                logger=_log,
                level="warning",
                cache_name=cache_path.name,
                payload_key=payload_key,
                error_type=type(exc).__name__,
            )
            record_metric(
                "industry_chain_cache_write_failures",
                1,
                unit="count",
                tags={"payload_key": payload_key, "error_type": type(exc).__name__},
                logger=_log,
            )
            return

    def read_rows(self, source_path: str | Path | None = None) -> list[dict]:
        path = self._resolve_source_path(source_path)
        if not path.exists():
            raise FileNotFoundError(str(path))
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl，无法读取 AI产业链.xlsx") from exc

        workbook = load_workbook(str(path), data_only=True, read_only=True)
        try:
            raw_rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        return build_ai_industry_chain_rows(raw_rows)

    def load_cached_rows(self, source_path: str | Path | None = None) -> list[dict]:
        path = self._resolve_source_path(source_path)
        if not self.is_default_source(path) or not path.exists():
            return []
        cached = self._read_signature_cache(self.rows_cache_path, path, "rows")
        if not isinstance(cached, list):
            return []
        return [dict(row) for row in cached if isinstance(row, dict)]

    def load_cached_stock_codes(self, source_path: str | Path | None = None) -> set[str]:
        path = self._resolve_source_path(source_path)
        if not self.is_default_source(path) or not path.exists():
            return set()
        cached = self._read_signature_cache(self.codes_cache_path, path, "stock_codes")
        if not isinstance(cached, list):
            return set()
        return {code for value in cached if (code := normalize_ai_chain_code(value))}

    def load_cached_context_map(self, source_path: str | Path | None = None) -> dict[str, str]:
        path = self._resolve_source_path(source_path)
        if not self.is_default_source(path) or not path.exists():
            return {}
        cached = self._read_signature_cache(self.context_cache_path, path, "context_map")
        if not isinstance(cached, dict):
            return {}
        return {
            code: str(text)
            for raw_code, text in cached.items()
            if (code := normalize_ai_chain_code(raw_code)) and str(text or "").strip()
        }

    def write_caches(self, rows: list[dict], source_path: str | Path | None = None) -> None:
        path = self._resolve_source_path(source_path)
        if not self.is_default_source(path):
            return
        stock_codes = sorted({code for row in rows if (code := normalize_ai_chain_code(row.get("代码")))})
        context_map = build_ai_industry_chain_context_map(rows)
        self._write_signature_cache(self.rows_cache_path, path, "rows", rows)
        self._write_signature_cache(self.codes_cache_path, path, "stock_codes", stock_codes)
        self._write_signature_cache(self.context_cache_path, path, "context_map", context_map)


__all__ = [
    "AI_CHAIN_CANDIDATE_PATHS",
    "AI_CHAIN_CODES_CACHE_FILE",
    "AI_CHAIN_CONTEXT_CACHE_FILE",
    "AI_CHAIN_FILE",
    "AI_CHAIN_ROWS_CACHE_FILE",
    "IndustryChainRepository",
    "resolve_ai_chain_file",
]
