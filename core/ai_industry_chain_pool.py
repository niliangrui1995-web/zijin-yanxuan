# -*- coding: utf-8 -*-
"""Shared AI industry chain stock-pool reader."""

from __future__ import annotations

import json
import threading
from pathlib import Path
from typing import Iterable

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_PROJECT_PARENT = Path(__file__).resolve().parents[2]
AI_CHAIN_FILE = _PROJECT_PARENT / "产业链投研" / "AI产业链.xlsx"
AI_CHAIN_ROWS_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_rows.json"
AI_CHAIN_CODES_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_stock_codes.json"
AI_CHAIN_CONTEXT_CACHE_FILE = _PROJECT_ROOT / "data" / "Cache" / "ai_industry_chain_context_map.json"
PLACEHOLDER = "--"
_CACHE_LOCK = threading.RLock()


def _cache_source_path(workbook_path: str | Path | None) -> Path:
    return Path(workbook_path) if workbook_path is not None else AI_CHAIN_FILE


def _is_default_cache_source(path: Path) -> bool:
    try:
        return path.resolve() == Path(AI_CHAIN_FILE).resolve()
    except OSError:
        return path == Path(AI_CHAIN_FILE)


def _source_signature(path: Path) -> dict:
    stat = path.stat()
    try:
        resolved = path.resolve()
    except OSError:
        resolved = path
    return {
        "path": str(resolved),
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def _read_signature_cache(cache_file: Path, source_path: Path, payload_key: str):
    signature = _source_signature(source_path)
    with _CACHE_LOCK:
        try:
            with cache_file.open("r", encoding="utf-8") as handle:
                payload = json.load(handle)
        except (OSError, TypeError, ValueError, json.JSONDecodeError):
            return None
    if not isinstance(payload, dict):
        return None
    if payload.get("source_signature") != signature:
        return None
    return payload.get(payload_key)


def _write_signature_cache(cache_file: Path, source_path: Path, payload_key: str, value) -> None:
    payload = {
        "source_signature": _source_signature(source_path),
        payload_key: value,
    }
    with _CACHE_LOCK:
        try:
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            with cache_file.open("w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=False)
        except (OSError, TypeError, ValueError):
            return


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_ai_chain_code(value) -> str:
    text = cell_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit() and len(text) <= 6:
        text = text.zfill(6)
    return text if len(text) == 6 and text.isdigit() else ""


def load_ai_industry_chain_rows(workbook_path: str | Path | None = None) -> list[dict]:
    path = Path(workbook_path) if workbook_path is not None else AI_CHAIN_FILE
    if not path.exists():
        raise FileNotFoundError(str(path))

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取 AI产业链.xlsx") from exc

    workbook = load_workbook(str(path), data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    headers = [cell_text(value) for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers) if header}

    def _get(row, header):
        idx = header_map.get(header)
        if idx is None or idx >= len(row):
            return ""
        return cell_text(row[idx])

    result = []
    for raw_row in rows[1:]:
        code = normalize_ai_chain_code(_get(raw_row, "代码"))
        if not code:
            continue
        result.append(
            {
                "代码": code,
                "名称": _get(raw_row, "公司名称") or _get(raw_row, "名称") or code,
                "现价": PLACEHOLDER,
                "涨幅": PLACEHOLDER,
                "市值": PLACEHOLDER,
                "细分板块": _get(raw_row, "细分板块") or _get(raw_row, "细分环节"),
                "5日涨幅": PLACEHOLDER,
                "10日涨幅": PLACEHOLDER,
                "20日涨幅": PLACEHOLDER,
                "备注": _get(raw_row, "备注"),
            }
        )
    return result


def load_cached_ai_industry_chain_rows(workbook_path: str | Path | None = None) -> list[dict]:
    path = _cache_source_path(workbook_path)
    if not _is_default_cache_source(path) or not path.exists():
        return []
    cached = _read_signature_cache(AI_CHAIN_ROWS_CACHE_FILE, path, "rows")
    if not isinstance(cached, list):
        return []
    return [dict(row) for row in cached if isinstance(row, dict)]


def load_cached_ai_industry_chain_stock_codes(workbook_path: str | Path | None = None) -> set[str]:
    path = _cache_source_path(workbook_path)
    if not _is_default_cache_source(path) or not path.exists():
        return set()
    cached = _read_signature_cache(AI_CHAIN_CODES_CACHE_FILE, path, "stock_codes")
    if not isinstance(cached, list):
        return set()
    return {normalize_ai_chain_code(code) for code in cached if normalize_ai_chain_code(code)}


def load_ai_industry_chain_stock_codes(workbook_path: str | Path | None = None) -> set[str]:
    path = _cache_source_path(workbook_path)
    if _is_default_cache_source(path):
        cached = load_cached_ai_industry_chain_stock_codes(workbook_path)
        if cached:
            return cached

    rows = refresh_ai_industry_chain_rows(workbook_path)
    stock_codes = {row["代码"] for row in rows if row.get("代码")}
    return stock_codes


def format_ai_industry_chain_context(row: dict | None, *, placeholder: str = PLACEHOLDER) -> str:
    if not isinstance(row, dict):
        return placeholder

    parts = []
    for key in ("细分板块", "细分环节", "备注"):
        text = cell_text(row.get(key))
        if text and text not in parts:
            parts.append(text)
    return " | ".join(parts) if parts else placeholder


def _build_ai_industry_chain_context_map(rows: Iterable[dict]) -> dict[str, str]:
    grouped: dict[str, list[str]] = {}
    for row in rows:
        code = row.get("代码")
        text = format_ai_industry_chain_context(row)
        if not code or text == PLACEHOLDER:
            continue
        bucket = grouped.setdefault(code, [])
        if text not in bucket:
            bucket.append(text)
    return {code: "；".join(values) for code, values in grouped.items()}


def load_cached_ai_industry_chain_context_map(workbook_path: str | Path | None = None) -> dict[str, str]:
    path = _cache_source_path(workbook_path)
    if not _is_default_cache_source(path) or not path.exists():
        return {}
    cached = _read_signature_cache(AI_CHAIN_CONTEXT_CACHE_FILE, path, "context_map")
    if not isinstance(cached, dict):
        return {}
    return {
        normalize_ai_chain_code(code): str(text)
        for code, text in cached.items()
        if normalize_ai_chain_code(code) and str(text or "").strip()
    }


def refresh_ai_industry_chain_rows(workbook_path: str | Path | None = None) -> list[dict]:
    path = _cache_source_path(workbook_path)
    rows = load_ai_industry_chain_rows(path)
    if _is_default_cache_source(path):
        stock_codes = sorted({row["代码"] for row in rows if row.get("代码")})
        context_map = _build_ai_industry_chain_context_map(rows)
        _write_signature_cache(AI_CHAIN_ROWS_CACHE_FILE, path, "rows", rows)
        _write_signature_cache(AI_CHAIN_CODES_CACHE_FILE, path, "stock_codes", stock_codes)
        _write_signature_cache(AI_CHAIN_CONTEXT_CACHE_FILE, path, "context_map", context_map)
    return rows


def load_ai_industry_chain_context_map(workbook_path: str | Path | None = None) -> dict[str, str]:
    path = _cache_source_path(workbook_path)
    if _is_default_cache_source(path):
        cached = load_cached_ai_industry_chain_context_map(workbook_path)
        if cached:
            return cached
    return _build_ai_industry_chain_context_map(refresh_ai_industry_chain_rows(workbook_path))


def normalize_stock_code_from_row(row: dict, code_keys: Iterable[str]) -> str:
    if not isinstance(row, dict):
        return ""
    for key in code_keys:
        code = normalize_ai_chain_code(row.get(key))
        if code:
            return code
    return ""


def filter_rows_to_ai_chain_codes(
    rows: Iterable[dict] | None,
    *,
    code_keys: Iterable[str] = ("代码", "股票代码", "证券代码", "stock_code"),
    stock_codes: Iterable[str] | None = None,
    workbook_path: str | Path | None = None,
) -> list[dict]:
    source_rows = list(rows or [])
    allowed_codes = (
        {normalize_ai_chain_code(code) for code in stock_codes}
        if stock_codes is not None
        else load_ai_industry_chain_stock_codes(workbook_path)
    )
    allowed_codes = {code for code in allowed_codes if code}
    return [row for row in source_rows if normalize_stock_code_from_row(row, code_keys) in allowed_codes]
